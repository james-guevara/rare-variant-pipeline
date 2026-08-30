#!/usr/bin/env -S uv run --script
# /// script
# requires-python = ">=3.11"
# dependencies = ["beautifulsoup4>=4.13", "openpyxl>=3.1"]
# ///
"""Build versioned candidate gene sets and overlap tables from downloaded sources."""

import argparse
import csv
import gzip
import json
import math
import re
from collections import defaultdict
from pathlib import Path

from bs4 import BeautifulSoup
from openpyxl import load_workbook

RETRIEVED = "2026-08-29"


def fnum(value):
    try:
        return float(value)
    except (TypeError, ValueError):
        return math.nan


def bh_selected(records, pfield, alpha=0.05):
    ranked = sorted(
        ((fnum(r.get(pfield)), r) for r in records if math.isfinite(fnum(r.get(pfield)))),
        key=lambda item: item[0],
    )
    cutoff = -1
    for i, (pvalue, _) in enumerate(ranked, 1):
        if pvalue <= alpha * i / len(ranked):
            cutoff = i
    return {id(r) for _, r in ranked[:cutoff]} if cutoff > 0 else set()


def write_tsv(path, rows, fields):
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="") as handle:
        writer = csv.DictWriter(
            handle, fieldnames=fields, delimiter="\t", extrasaction="ignore", lineterminator="\n"
        )
        writer.writeheader()
        writer.writerows({field: "." if row.get(field) in (None, "") else row.get(field) for field in fields}
                         for row in rows)


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--raw", type=Path, default=Path("resources/gene-sets/raw/2026-08-29"))
    parser.add_argument("--output", type=Path, default=Path("resources/gene-sets/processed/2026-08-29"))
    args = parser.parse_args()
    raw, output = args.raw, args.output
    members = []
    sources = []

    def source(source_id, title, release, url, restriction=""):
        sources.append(dict(source_id=source_id, title=title, release=release, url=url,
                            retrieved_date=RETRIEVED, usage_note=restriction))

    def add(set_id, phenotype, definition, symbol, ensembl="", evidence="", direction="",
            beta="", pvalue="", source_id="", notes=""):
        symbol = (symbol or "").strip().upper()
        if not symbol:
            return
        members.append(dict(gene_set_id=set_id, phenotype=phenotype, definition=definition,
                            gene_symbol=symbol, ensembl_gene_id=ensembl, evidence=evidence,
                            effect_direction=direction, effect_beta=beta, p_value=pvalue,
                            source_id=source_id, notes=notes))

    # SFARI: preserve all scored categories and syndromic status as separate candidate definitions.
    sfari_file = raw / "sfari-gene-scoring-2026Q2.csv"
    source("sfari_2026q2", "SFARI Gene scoring", "2026 Q2",
           "https://gene-development.sfari.org/database/gene-scoring/")
    with sfari_file.open(newline="") as handle:
        for row in csv.DictReader(handle):
            symbol, score = row["gene-symbol"], row["gene-score"]
            if score:
                add("sfari_all_scored", "ASD", "All genes assigned a SFARI score", symbol,
                    row["ensembl-id"], score, source_id="sfari_2026q2")
                add(f"sfari_score_{score}", "ASD", f"SFARI score {score}", symbol,
                    row["ensembl-id"], score, source_id="sfari_2026q2")
            if row["syndromic"] == "1":
                add("sfari_syndromic", "ASD", "SFARI syndromic flag", symbol,
                    row["ensembl-id"], score, source_id="sfari_2026q2")

    # PanelApp: retain broad panel membership and high-evidence (green) definitions.
    panel_specs = [
        ("panelapp_intellectual_disability", "Intellectual disability", 285, "11.26",
         "panelapp-intellectual-disability-panel-285-v11.26.json"),
        ("panelapp_epilepsy", "Epilepsy", 402, "9.73",
         "panelapp-early-onset-or-syndromic-epilepsy-panel-402-v9.73.json"),
    ]
    for prefix, phenotype, panel_id, version, filename in panel_specs:
        sid = f"{prefix}_v{version}"
        source(sid, f"PanelApp panel {panel_id}: {phenotype}", version,
               f"https://panelapp.genomicsengland.co.uk/panels/{panel_id}/")
        payload = json.loads((raw / filename).read_text())
        for item in payload["genes"]:
            symbol = item.get("entity_name") or item.get("gene_data", {}).get("gene_symbol")
            grch38 = item.get("gene_data", {}).get("ensembl_genes", {}).get("GRch38", {})
            if isinstance(grch38, dict):
                ensembl = next((v.get("ensembl_id", "") for v in grch38.values() if isinstance(v, dict)), "")
            else:
                ensembl = grch38
            confidence = str(item.get("confidence_level", ""))
            add(prefix + "_all", phenotype, "All genes in the versioned PanelApp panel",
                symbol, ensembl, confidence, source_id=sid)
            if confidence == "3":
                add(prefix + "_green", phenotype, "PanelApp green/high-evidence genes",
                    symbol, ensembl, confidence, source_id=sid)

    # Published 2022 SCHEMA exome-wide table (the historical SCHEMA-10 comparator).
    source("schema_2022", "Published SCHEMA study", "2022", "https://pmc.ncbi.nlm.nih.gov/articles/PMC9805802/")
    soup = BeautifulSoup((raw / "schema-published-2022-article.html").read_text(), "html.parser")
    for tr in soup.find("table").find_all("tr")[1:]:
        cells = [x.get_text(" ", strip=True) for x in tr.find_all(["th", "td"])]
        if cells:
            add("schema_2022_exomewide_10", "Schizophrenia",
                "Ten published exome-wide significant SCHEMA genes", cells[0], evidence="exome-wide",
                pvalue=cells[10], source_id="schema_2022", notes=f"Q={cells[11]}")

    # Phase II SCHEMA snapshots. Use the final release as the ENSG-to-symbol map for early snapshots.
    schema_files = sorted(raw.glob("schema-phase2-*-gene-results.tsv.bgz"))
    symbol_map = {}
    with gzip.open(schema_files[-1], "rt") as handle:
        for row in csv.DictReader(handle, delimiter="\t"):
            symbol_map[row["gene_id"]] = row.get("gene_symbol", "")
    for path in schema_files:
        release = re.search(r"(2026-\d\d-\d\d)", path.name).group(1)
        sid = f"schema_phase2_{release}"
        source(sid, "SCHEMA Phase II gene results", release, "https://schema.broadinstitute.org/",
               "Current browser terms restrict some global/exome-wide and large gene-set analyses before the updated paper.")
        with gzip.open(path, "rt") as handle:
            rows = [r for r in csv.DictReader(handle, delimiter="\t") if r.get("group") == "meta"]
        pfield = "case_control_plus_de_novo_p_value"
        finite = [r for r in rows if math.isfinite(fnum(r.get(pfield)))]
        bonf = 0.05 / len(finite)
        fdr_ids = bh_selected(finite, pfield)
        for row in finite:
            symbol = row.get("gene_symbol") or symbol_map.get(row["gene_id"], "")
            pvalue = fnum(row[pfield])
            if pvalue <= bonf:
                add(f"schema_phase2_{release}_bonferroni", "Schizophrenia",
                    f"Phase II meta-analysis P <= 0.05/{len(finite)}", symbol, row["gene_id"],
                    "Bonferroni", pvalue=pvalue, source_id=sid)
            if id(row) in fdr_ids:
                add(f"schema_phase2_{release}_fdr05", "Schizophrenia",
                    "Phase II meta-analysis Benjamini-Hochberg FDR <= 5%", symbol, row["gene_id"],
                    "FDR<=0.05", pvalue=pvalue, source_id=sid)

    # Epi25: preserve all phenotype groups, consequence models, and two significance definitions.
    source("epi25_2022", "Epi25 gene results", "2022-12-01",
           "https://epi25.broadinstitute.org/")
    with gzip.open(raw / "epi25-2022-12-01-gene-results.tsv.bgz", "rt") as handle:
        epi_rows = list(csv.DictReader(handle, delimiter="\t"))
    for group in sorted({r["group"] for r in epi_rows}):
        group_rows = [r for r in epi_rows if r["group"] == group]
        for model, pfield in [("ptv", "ptv_pval"), ("damaging_missense", "damaging_missense_pval")]:
            finite = [r for r in group_rows if math.isfinite(fnum(r.get(pfield)))]
            fdr_ids = bh_selected(finite, pfield)
            bonf = 0.05 / len(finite)
            for row in finite:
                pvalue = fnum(row[pfield])
                stem = f"epi25_{group.lower()}_{model}"
                if pvalue <= bonf:
                    add(stem + "_bonferroni", "Epilepsy", f"Epi25 {group} {model}; Bonferroni",
                        row.get("gene_symbol") or symbol_map.get(row["gene_id"], ""), row["gene_id"], "Bonferroni", pvalue=pvalue,
                        source_id="epi25_2022")
                if id(row) in fdr_ids:
                    add(stem + "_fdr05", "Epilepsy", f"Epi25 {group} {model}; BH FDR <= 5%",
                        row.get("gene_symbol") or symbol_map.get(row["gene_id"], ""), row["gene_id"], "FDR<=0.05", pvalue=pvalue,
                        source_id="epi25_2022")

    # Height: comprehensive 207-gene set plus signed singleton-pLoF subsets requested by the PI.
    source("kosmicki_height_2026", "Kosmicki et al. height preprint supplement", "2026-06-22",
           "https://www.medrxiv.org/content/10.64898/2026.06.22.26355163v1")
    workbook = load_workbook(raw / "kosmicki-height-2026-supplementary-tables.xlsx", read_only=True, data_only=True)
    for sheet_name, set_id, definition in [
        ("Table S4", "height_207_gene_p", "207 discovery-significant gene-P genes"),
        ("Table S5", "height_singleton_plof", "17 significant singleton-pLoF burden genes"),
    ]:
        sheet = workbook[sheet_name]
        headers = [str(x.value) if x.value is not None else "" for x in next(sheet.iter_rows(min_row=2, max_row=2))]
        for values in sheet.iter_rows(min_row=3, values_only=True):
            row = dict(zip(headers, values))
            beta = row.get("Effect (SD units) (discovery+replication)", "")
            direction = "positive" if fnum(beta) > 0 else "negative" if fnum(beta) < 0 else ""
            add(set_id, "Height", definition, row.get("Gene symbol"), row.get("Ensembl ID", ""),
                direction=direction, beta=beta,
                pvalue=row.get("P-value (discovery+replication)", row.get("best P-value (discovery)", "")),
                source_id="kosmicki_height_2026")
            if sheet_name == "Table S5" and direction:
                add(f"height_singleton_plof_{direction}", "Height", f"Singleton-pLoF genes with {direction} beta",
                    row.get("Gene symbol"), row.get("Ensembl ID", ""), direction=direction, beta=beta,
                    pvalue=row.get("P-value (discovery+replication)", ""), source_id="kosmicki_height_2026")

    # BMI: signed 16-gene exome-wide table.
    source("bmi_exome_2021", "Exome-wide BMI burden associations", "2021",
           "https://pmc.ncbi.nlm.nih.gov/articles/PMC10275396/")
    soup = BeautifulSoup((raw / "bmi-exome-2021-article.html").read_text(), "html.parser")
    for tr in soup.find("table").find_all("tr")[1:]:
        cells = [x.get_text(" ", strip=True) for x in tr.find_all(["th", "td"])]
        if not cells:
            continue
        beta_text = cells[3].replace("−", "-").strip()
        beta = fnum(re.match(r"-?\s*\d+(?:\.\d+)?", beta_text).group().replace(" ", ""))
        direction = "positive" if beta > 0 else "negative"
        add("bmi_exomewide_16", "BMI", "Sixteen exome-wide significant BMI burden genes", cells[0],
            evidence="exome-wide", direction=direction, beta=beta, pvalue=cells[4],
            source_id="bmi_exome_2021", notes=cells[2])
        add(f"bmi_exomewide_{direction}", "BMI", f"Exome-wide BMI genes with {direction} beta", cells[0],
            evidence="exome-wide", direction=direction, beta=beta, pvalue=cells[4],
            source_id="bmi_exome_2021", notes=cells[2])

    # Deduplicate exact set/gene memberships, then calculate summaries and pairwise overlaps.
    unique = {(r["gene_set_id"], r["gene_symbol"]): r for r in members}
    members = sorted(unique.values(), key=lambda r: (r["gene_set_id"], r["gene_symbol"]))
    sets = defaultdict(set)
    definitions = {}
    phenotypes = {}
    for row in members:
        sets[row["gene_set_id"]].add(row["gene_symbol"])
        definitions[row["gene_set_id"]] = row["definition"]
        phenotypes[row["gene_set_id"]] = row["phenotype"]
    summary = [dict(gene_set_id=s, phenotype=phenotypes[s], n_genes=len(genes), definition=definitions[s])
               for s, genes in sorted(sets.items())]
    overlaps = []
    for left in sorted(sets):
        for right in sorted(sets):
            intersection = len(sets[left] & sets[right])
            union = len(sets[left] | sets[right])
            overlaps.append(dict(left_gene_set=left, right_gene_set=right,
                                 left_n=len(sets[left]), right_n=len(sets[right]),
                                 intersection_n=intersection, union_n=union,
                                 jaccard=intersection / union if union else 0))

    schema_comparisons = []
    phase2_bonf = sorted(s for s in sets if s.startswith("schema_phase2_") and s.endswith("_bonferroni"))
    comparison_pairs = [("schema_2022_exomewide_10", phase2_bonf[-1])]
    comparison_pairs += list(zip(phase2_bonf, phase2_bonf[1:]))
    for left, right in comparison_pairs:
        for gene in sorted(sets[left] | sets[right]):
            status = "shared" if gene in sets[left] and gene in sets[right] else "gained" if gene in sets[right] else "lost"
            schema_comparisons.append(dict(comparison_id=f"{left}__vs__{right}", left_gene_set=left,
                                           right_gene_set=right, gene_symbol=gene, status=status))

    write_tsv(output / "gene_set_membership.tsv", members,
              ["gene_set_id", "phenotype", "definition", "gene_symbol", "ensembl_gene_id", "evidence",
               "effect_direction", "effect_beta", "p_value", "source_id", "notes"])
    write_tsv(output / "gene_set_summary.tsv", summary,
              ["gene_set_id", "phenotype", "n_genes", "definition"])
    write_tsv(output / "gene_set_overlap.tsv", overlaps,
              ["left_gene_set", "right_gene_set", "left_n", "right_n", "intersection_n", "union_n", "jaccard"])
    write_tsv(output / "gene_set_sources.tsv", sources,
              ["source_id", "title", "release", "url", "retrieved_date", "usage_note"])
    write_tsv(output / "schema_release_comparison.tsv", schema_comparisons,
              ["comparison_id", "left_gene_set", "right_gene_set", "gene_symbol", "status"])
    print(f"Wrote {len(members):,} memberships across {len(sets)} nonempty gene sets to {output}")


if __name__ == "__main__":
    main()
